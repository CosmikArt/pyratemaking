"""Excel export for rate-filing exhibits.

Each DataFrame in ``sheets`` becomes a worksheet. Numeric columns get a
default number format; user-supplied formats override.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


_HEADER_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
_HEADER_FONT = Font(bold=True)


def write_excel(
    path: str | Path,
    sheets: dict[str, pd.DataFrame],
    *,
    column_formats: dict[str, dict[str, str]] | None = None,
) -> Path:
    """Write multiple DataFrames to a single .xlsx file.

    Parameters
    ----------
    path : str or Path
    sheets : dict[str, DataFrame]
        Sheet name → frame.
    column_formats : dict, optional
        ``{sheet_name: {column_name: number_format}}`` to override defaults.

    Returns
    -------
    Path of the written file.
    """
    column_formats = column_formats or {}
    out = Path(path)
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        rows = list(dataframe_to_rows(df.reset_index(), index=False, header=True))
        for row in rows:
            ws.append(row)
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left")

        formats = column_formats.get(sheet_name, {})
        for col_idx, header_cell in enumerate(ws[1], start=1):
            col_letter = header_cell.column_letter
            fmt = formats.get(str(header_cell.value))
            if fmt is None:
                fmt = _default_format(df, str(header_cell.value))
            if fmt:
                for cell in ws[col_letter][1:]:
                    cell.number_format = fmt
    wb.save(out)
    return out


def _default_format(df: pd.DataFrame, column_name: str) -> str | None:
    if column_name not in df.columns:
        return None
    series = df[column_name]
    if not pd.api.types.is_numeric_dtype(series):
        return None
    name = column_name.lower()
    if "premium" in name or "loss" in name or "amount" in name:
        return '"$"#,##0'
    if "factor" in name or "relativity" in name or "ldf" in name or "cdf" in name:
        return "0.0000"
    if "ratio" in name or "pct" in name or "change" in name:
        return "0.00%"
    return "#,##0.00"
