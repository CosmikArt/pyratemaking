from pathlib import Path

import pandas as pd
import pytest

from pyratemaking.reporting import (
    RatePlanReport,
    format_currency,
    format_percent,
    format_relativity,
    style_actuarial_table,
    write_excel,
)
from pyratemaking.reporting.filing import render_filing_html


def test_format_helpers_handle_nan():
    assert format_currency(float("nan")) == "—"
    assert format_percent(float("nan")) == "—"
    assert format_relativity(float("nan")) == "—"


def test_format_helpers_known_values():
    assert format_currency(1234567.89) == "$1,234,568"
    assert format_percent(0.05).startswith("+5.00")
    assert format_relativity(0.95) == "0.9500"


def test_style_actuarial_table_renders_html():
    df = pd.DataFrame({"premium": [1000.0], "relativity": [0.85]})
    styler = style_actuarial_table(
        df, currency_cols=["premium"], relativity_cols=["relativity"], title="Exhibit"
    )
    html = styler.to_html()
    assert "Exhibit" in html
    assert "$1,000" in html
    assert "0.8500" in html


def test_filing_html_renders_with_minimal_inputs():
    report = RatePlanReport(title="Test filing")
    html = render_filing_html(report)
    assert "Test filing" in html
    assert "<html" in html


def test_filing_html_includes_indication_when_present():
    report = RatePlanReport(
        title="Filing",
        indication={"rate_change": 0.05, "method": "loss_ratio", "credibility": 1.0},
    )
    html = render_filing_html(report)
    assert "+5.00%" in html


def test_rateplan_report_writes_to_disk(tmp_path: Path):
    report = RatePlanReport(title="Filing")
    out = report.write(tmp_path / "filing.html")
    assert out.exists()
    assert "<html" in out.read_text()


def test_write_excel_creates_file_with_correct_sheets(tmp_path: Path):
    sheets = {
        "indication": pd.DataFrame({"rate_change": [0.05], "premium": [1_000_000]}),
        "onleveling": pd.DataFrame({"on_level_factor": [1.05], "earned_premium": [950_000]}),
    }
    out = write_excel(tmp_path / "filing.xlsx", sheets)
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"indication", "onleveling"}
